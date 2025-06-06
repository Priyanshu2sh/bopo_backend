from datetime import date, datetime, timezone
from io import BytesIO
import json
import os
import random
import string
from tkinter.font import Font
from django.db.models import Max
from django.http import HttpResponse, JsonResponse
from django.db.models.functions import Cast, Substr
from django.shortcuts import get_object_or_404, render
import openpyxl
from requests import Response
from rest_framework import status
from openpyxl.styles import Font
from twilio.rest import Client
from django.conf import settings
from django.contrib.auth.hashers import check_password



from accounts import models
from accounts.models import Corporate, Customer, Merchant, Terminal
from accounts.views import generate_terminal_id
from accounts.models import Corporate, Customer, Merchant, Terminal
from accounts.views import generate_terminal_id
from bopo_award.models import CustomerPoints, History, MerchantPoints, PaymentDetails

# from django.contrib.auth import authenticate 
# from django.shortcuts import redirect
from .models import AccountInfo, BopoAdmin, DeductSetting, Employee, EmployeeRole, MerchantCredential, MerchantLogin, Notification, Reducelimit, SecurityQuestion, Topup, UploadedFile
from django.http import JsonResponse
from .models import State, City
from bopo_admin.models import Employee
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction




# Create your views here.
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm

# def login(request):
#     if request.method == 'POST':
#         form = AuthenticationForm(data=request.POST)
#         if form.is_valid():
#             username = form.cleaned_data['username']
#             password = form.cleaned_data['password']
#             user = authenticate(username=username, password=password)
#             if user is not None:
#                 auth_login(request, user)
#                 return redirect('home')  # Redirect to home after successful login
#             else:
#                 error_message = 'Invalid login credentials'
#                 return render(request, 'login.html', {'error_message': error_message})
#     else:
#         form = AuthenticationForm()
#     return render(request, 'login.html', {'form': form})

from django.contrib.auth import logout
def custom_logout_view(request):
    logout(request)
    return redirect('login')

def profile(request):
    return render(request, 'bopo_admin/profile.html')  # Include the correct path and .html extension


def corporate_admin(request):
    corporates = Corporate.objects.all()
    corporate_data = []

    for corporate in corporates:
        # Fetch merchants linked to the corporate
        merchants = Merchant.objects.filter(corporate_id=corporate.corporate_id, user_type='corporate')
        corporate_data.append({
            "corporate": corporate,
            "merchants": merchants
        })

    return render(request, 'bopo_admin/Payment/corporate_admin.html', {
        "corporate_data": corporate_data
    })


# def dashboard(request):
#     print("User role in session:", request.session.get('user_role'))  # Debug line
#     user_role = request.session.get('user_role')
#     return render(request, 'base.html', {'user_role': user_role})


from django.shortcuts import render
from accounts.models import Merchant  # Replace `your_app` and `Merchant` with actual names

def terminals(request):
    merchants = Merchant.objects.all().order_by('merchant_id')
    return render(request, 'bopo_admin/Payment/terminals.html', {'merchants': merchants})

from django.http import JsonResponse
from accounts.models import Merchant, Terminal

def get_terminals(request, merchant_id):
    try:
        merchant = Merchant.objects.get(merchant_id=merchant_id)
    except Merchant.DoesNotExist:
        return JsonResponse({'error': 'Merchant not found'}, status=404)

    terminals = Terminal.objects.filter(merchant_id=merchant)

    terminal_data = [
        {'terminal_id': terminal.terminal_id, 'tid_pin': terminal.tid_pin}
        for terminal in terminals
    ]

    return JsonResponse({'terminals': terminal_data})

import random
import string
from django.http import JsonResponse
from accounts.models import Merchant, Terminal

def generate_terminal_id():
    """Generate a unique terminal ID."""
    return "TID" + ''.join(random.choices(string.digits, k=8))

def add_terminal(request, merchant_id):
    """Generate a new terminal and pin for the merchant."""
    try:
        merchant = Merchant.objects.get(merchant_id=merchant_id)
    except Merchant.DoesNotExist:
        return JsonResponse({'error': 'Merchant not found'}, status=404)

    # Generate unique terminal ID
    terminal_id = generate_terminal_id()
    while Terminal.objects.filter(terminal_id=terminal_id).exists():
        terminal_id = generate_terminal_id()  # Ensure it's unique

    # Generate a 4-digit PIN
    tid_pin = random.randint(1000, 9999)

    # Save terminal info to the database
    terminal = Terminal.objects.create(
        terminal_id=terminal_id,
        tid_pin=tid_pin,
        merchant_id=merchant
    )

    # Return the newly created terminal details
    return JsonResponse({'terminal_id': terminal.terminal_id, 'tid_pin': terminal.tid_pin})

# In views.py

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

@csrf_exempt
def update_terminal_pin(request, merchant_id, terminal_id):
    if request.method == 'POST':
        body = json.loads(request.body)
        new_pin = body.get('tid_pin')

        try:
            # Fetch the merchant by its ID
            merchant = Merchant.objects.get(merchant_id=merchant_id)
            
            # Now fetch the terminal by both merchant and terminal_id
            terminal = Terminal.objects.get(merchant_id=merchant, terminal_id=terminal_id)
            
            terminal.tid_pin = new_pin
            terminal.save()
            return JsonResponse({'success': True})
        except Merchant.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Merchant not found'})
        except Terminal.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Terminal not found'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})

import json
from django.http import JsonResponse
from datetime import datetime
from accounts.models import Terminal

def toggle_terminal_status(request, terminal_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            is_active = data.get("is_active")

            terminal = Terminal.objects.get(id=terminal_id)

            # Change terminal status based on the checkbox state
            if is_active:
                terminal.status = "Active"
            else:
                terminal.status = "Inactive"

            terminal.save()

            return JsonResponse({"success": True, "status": terminal.status})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request"})

 
@login_required
def home(request):
    user = request.user

    # Calculate total projects and project progress
    total_projects = Corporate.objects.count()
    completed_projects = Corporate.objects.filter(status="Completed").count()  # Assuming a 'status' field exists
    project_progress = (completed_projects / total_projects * 100) if total_projects > 0 else 0

    # Calculate total merchants and merchant progress
    total_merchants = Merchant.objects.count()
    active_merchants = Merchant.objects.filter(status="Active").count()  # Assuming a 'status' field exists
    merchant_progress = (active_merchants / total_merchants * 100) if total_merchants > 0 else 0

    # Calculate total customers
    total_customers = Customer.objects.count()

    # Calculate total users (sum of customers, merchants, and corporates)
    total_users = total_customers + total_merchants + total_projects

     # Prepare data for the bar chart
    chart_data = {
        "projects": [total_projects, completed_projects],
        "merchants": [total_merchants, active_merchants],
        "customers": [total_customers],
    }

    # Add context for the dashboard
    context = {
        "user": user,
        "title": "Welcome to Bopo Admin Dashboard",
        "description": "Manage merchants, customers, and projects efficiently.",
        "total_projects": total_projects,
        "project_progress": project_progress,
        "total_merchants": total_merchants,
        "merchant_progress": merchant_progress,
        "total_customers": total_customers,
        "total_users": total_users,
        "chart_data": chart_data,
    }
    if user.role == 'corporate_admin':
        return render(request, 'bopo_admin/Corporate/corporate_dashboard.html', context)
    elif user.role == 'employee':
        employee = user.employee
        role_permissions = EmployeeRole.objects.get(employee=employee)
        context['role_permissions'] = role_permissions
        if role_permissions.corporate_merchant or role_permissions.individual_merchant or role_permissions.merchant_send_credentials or role_permissions.merchant_limit or role_permissions.merchant_login_page_info or role_permissions.merchant_send_notification or role_permissions.merchant_received_offers:
            context['merchant'] = True
        return render(request, 'bopo_admin/home.html', context)
    else:
        return render(request, 'bopo_admin/home.html', context)

def about(request):
     return render(request, 'bopo_admin/about.html')

def merchant(request):
    return render(request, 'bopo_admin/Merchant/merchant.html')

def customer(request):
    customers = Customer.objects.all()
    return render(request, 'bopo_admin/Customer/customer.html', {'customers': customers})


from django.http import JsonResponse, Http404
from accounts.models import Customer  # Adjust this import to your actual model
from django.shortcuts import get_object_or_404

def get_customer(request, customer_id):
    customer = get_object_or_404(Customer, customer_id=customer_id)

    # Retrieve the state object by its name (if state is stored as a string)
    state_obj = State.objects.get(name=customer.state)  # Assuming state is a string, get State object by name

    # Retrieve cities based on selected state
    cities = City.objects.filter(state=state_obj)  # Now we use the State object

    # Convert cities to a dictionary for use in the frontend
    city_data = [{"id": city.id, "name": city.name} for city in cities]

    # Data to send to the frontend
    data = {
        "first_name": customer.first_name,
        "last_name": customer.last_name,
        "email": customer.email,
        "mobile": customer.mobile,
        "age": customer.age,
        "aadhar_number": customer.aadhar_number,
        "pin": customer.pin,
        "address": customer.address,
        "pincode": customer.pincode,
        "gender": customer.gender,
        "pan_number": customer.pan_number,
        "state": customer.state,  # Prefilled state (assuming it's a string or related field)
        "city": customer.city,    # Prefilled city (assuming it's a string or related field)
        "states": [{"id": state.id, "name": state.name} for state in State.objects.all()],  # List of all states
        "cities": city_data,  # List of cities filtered by the selected state
    }

    return JsonResponse(data)

from django.shortcuts import render, get_object_or_404, redirect
from accounts.models import Customer
from django.http import JsonResponse


def update_customer(request, customer_id):  # <-- accept customer_id here
    if request.method == "POST":
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        mobile = request.POST.get('mobile')
        age = request.POST.get('age')
        aadhar_number = request.POST.get('aadhar_number')
        pin = request.POST.get('pin')
        address = request.POST.get('address')
        state_id = request.POST.get('state')
        city_id = request.POST.get('city')
        pincode = request.POST.get('pincode')
        gender = request.POST.get('gender')
        pan_number = request.POST.get('pan_number')

        try:
            customer = Customer.objects.get(customer_id=customer_id)  # Note: use `customer_id` field

            customer.first_name = first_name
            customer.last_name = last_name
            customer.email = email
            customer.mobile = mobile
            customer.age = age
            customer.aadhar_number = aadhar_number
            customer.pin = pin
            customer.address = address
            customer.pincode = pincode
            customer.gender = gender
            customer.pan_number = pan_number

            if state_id:
                state_obj = State.objects.get(id=state_id)
                customer.state = state_obj.name  # or assign FK if applicable
            if city_id:
                city_obj = City.objects.get(id=city_id)
                customer.city = city_obj.name  # or assign FK if applicable

            customer.save()

            return JsonResponse({
                "success": True,
                "message": "Customer updated successfully!"
            })

        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Customer not found'})
        except State.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'State not found'})
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'City not found'})

    return JsonResponse({'success': False, 'error': 'Invalid request'})



# bopo_admin/views.py

from django.http import JsonResponse
from accounts.models import Customer


def delete_customer(request, customer_id):
    if request.method == 'DELETE':
        try:
            customer = Customer.objects.get(customer_id=customer_id)
            customer.delete()
            return JsonResponse({'status': 'success', 'message': 'Customer deleted successfully.'})
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Customer not found.'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)


def merchant_list(request):
    return render(request, "bopo_admin/Merchant/merchant_list.html")

def corporate_list(request):
    corporates = Corporate.objects.all()
    corporate_data = []

    for corporate in corporates:
        # Fetch merchants linked to the corporate
        merchants = Merchant.objects.filter(corporate_id=corporate.corporate_id, user_type='corporate')
        corporate_data.append({
            "corporate": corporate,
            "merchants": merchants
        })

    return render(request, 'bopo_admin/Merchant/corporate_list.html', {
        "corporate_data": corporate_data
    })

def individual_list(request):
    merchants = Merchant.objects.filter(user_type='individual')  # Fetch only individual merchants
    merchant_points = MerchantPoints.objects.filter(merchant__in=merchants)  # Fetch points for these merchants

    # Create a mapping of merchant IDs to their actual points
    points_mapping = {mp.merchant.id: mp.points for mp in merchant_points}

    # Pass merchants and their points to the template
    return render(request, "bopo_admin/Merchant/individual_list.html", {
        "merchants": merchants,
        "points_mapping": points_mapping
    })


def toggle_status(request, merchant_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            is_active = data.get("is_active")

            merchant = Merchant.objects.get(id=merchant_id)

            if is_active:
                merchant.status = "Active"
                merchant.verified_at = datetime.now()
            else:
                merchant.status = "Inactive"
                merchant.verified_at = None

            merchant.save()

            return JsonResponse({"success": True, "status": merchant.status})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request"})


from django.shortcuts import render, redirect
from django.contrib import messages
from accounts.models import Merchant, Corporate
from django.http import JsonResponse
import random
import string

def add_merchant(request):
    if request.method == "POST":
        try:
            is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

            select_project = request.POST.get("select_project")
            project_type = request.POST.get("project_type")
            project_name = request.POST.get("project_name", "")
            first_name = request.POST.get("first_name")
            last_name = request.POST.get("last_name")
            email = request.POST.get("email")
            mobile = request.POST.get("mobile")
            aadhaar_number = request.POST.get("aadhaar_number")
            pin = request.POST.get("pin")
            gst_number = request.POST.get("gst_number")
            shop_name = request.POST.get("shop_name")
            pan_number = request.POST.get("pan_number")
            address = request.POST.get("address")
            legal_name = request.POST.get("legal_name")
            pincode = request.POST.get("pincode")
            city_id = request.POST.get("city")
            state_id = request.POST.get("state")
            country = request.POST.get("country", "India")
            state = State.objects.get(id=state_id)
            city = City.objects.get(id=city_id)



            # ✅ Unique field checks
            if Merchant.objects.filter(email=email).exists() or Corporate.objects.filter(email=email).exists():
                message = "Email is already registered."
                return JsonResponse({"success": False, "message": message}) if is_ajax else redirect_with_error(message)

            if Merchant.objects.filter(mobile=mobile).exists() or Corporate.objects.filter(mobile=mobile).exists():
                message = "Mobile number is already registered."
                return JsonResponse({"success": False, "message": message}) if is_ajax else redirect_with_error(message)

            if Merchant.objects.filter(aadhaar_number=aadhaar_number).exists() or Corporate.objects.filter(aadhaar_number=aadhaar_number).exists():
                message = "Aadhaar number is already registered."
                return JsonResponse({"success": False, "message": message}) if is_ajax else redirect_with_error(message)

            # if Merchant.objects.filter(pan_number=pan).exists():
            #     message = "PAN number is already registered."
            #     return JsonResponse({"success": False, "message": message}) if is_ajax else redirect_with_error(message)

            # ✅ Corporate ID
            last_corporate = Corporate.objects.exclude(corporate_id=None).order_by("-corporate_id").first()
            new_corporate_id = 1 if not last_corporate else int(last_corporate.corporate_id[6:]) + 1
            corporate_id = f"CORP{new_corporate_id:06d}"

            if project_type == "Existing Project" and select_project:
                corporate = Corporate.objects.get(id=select_project)
                project_name = corporate.project_name
                project_id = corporate.project_id

                # ✅ Merchant ID
                project_abbr = project_name[:4].upper()
                random_number = ''.join(random.choices(string.digits, k=11))
                merchant_id = f"{project_abbr}{random_number}"

                Merchant.objects.create(
                    user_type='corporate',
                    merchant_id=merchant_id,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    mobile=mobile,
                    aadhaar_number=aadhaar_number,
                    pin=pin,
                    gst_number=gst_number,
                    pan_number=pan_number,
                    shop_name=shop_name,
                    legal_name=legal_name,
                    address=address,
                    pincode=pincode,
                    state=state,
                    city=city,
                    country=country,
                    corporate_id=corporate.corporate_id,
                    project_name=corporate
                    
                )

                merchant= Merchant.objects.get(merchant_id=merchant_id)
                
                
               # ✅ Generate Terminal ID and TID PIN
                def generate_terminal_id():
                    return "TID" + ''.join(random.choices(string.digits, k=8))

                terminal_id = generate_terminal_id()
                while Terminal.objects.filter(terminal_id=terminal_id).exists():
                    terminal_id = generate_terminal_id()

                tid_pin = random.randint(1000, 9999)

                # ✅ Save terminal info
                Terminal.objects.create(
                    terminal_id=terminal_id,
                    tid_pin=tid_pin,
                    merchant_id=merchant
                )

            elif project_type == "New Project":
                if not project_name:
                    message = "Project name is required for new projects."
                    return JsonResponse({"success": False, "message": message}) if is_ajax else redirect_with_error(message)

                last_project = Corporate.objects.exclude(project_id=None).order_by("-project_id").first()
                new_project_id = 1 if not last_project else int(last_project.project_id[4:]) + 1
                project_id = f"PROJ{new_project_id:06d}"

                corporate = Corporate.objects.create(
                    select_project=select_project,
                    corporate_id=corporate_id,
                    project_name=project_name,
                    project_id=project_id,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    mobile=mobile,
                    aadhaar_number=aadhaar_number,
                    pin=pin,
                    gst_number=gst_number,
                    pan_number=pan_number,
                    shop_name=shop_name,
                    legal_name=legal_name,
                    address=address,
                    pincode=pincode,
                    state=state,
                    city=city,
                    country=country,
                    role="admin",
                )

                bopo_admin = BopoAdmin(username=project_name, role="corporate_admin", corporate=corporate)
                bopo_admin.set_password(pin)  # Hash the password
                bopo_admin.save()

                try:
                    phone_number = corporate.mobile
                    if not phone_number.startswith('+'):
                        phone_number = f'+91{phone_number}'

                # Compose the SMS message
                    message_text = (
                        f"Dear {corporate.first_name},\n\n"
                        f"Your BOPO login credentials are as follows:\n"
                        f"Project Name : {corporate.project_name}\n\n"
                        f"Password : {corporate.pin}\n"
                        f"Please use these credentials to access your BOPO admin panel.\n\n"
                        f"Regards,\n"
                        f"BOPO Support Team"
                    )
                
                    # Fetch Twilio credentials from Django settings
                    account_sid = settings.TWILIO_ACCOUNT_SID
                    auth_token = settings.TWILIO_AUTH_TOKEN
                    twilio_phone_number = settings.TWILIO_PHONE_NUMBER

                    # Send SMS using Twilio
                    client = Client(account_sid, auth_token)
                    client.messages.create(
                        body=message_text,
                        from_=twilio_phone_number,
                        to=phone_number
                    )
                
                    messages.success(request, f"Credentials sent to {merchant.first_name} at {phone_number}")

                except Exception as e:
                    messages.error(request, f"Error sending SMS: {str(e)}")

            else:
                message = "Invalid project type selected."
                return JsonResponse({"success": False, "message": message}) if is_ajax else redirect_with_error(message)

            success_message = "Merchant added successfully."
            return JsonResponse({"success": True, "message": success_message}) if is_ajax else redirect_with_success(success_message)

        except Exception as e:
            print("Error saving merchant:", e)
            return JsonResponse({"success": False, "message": "Something went wrong. Please check your inputs."})

    corporates = Corporate.objects.all()
    return render(request, "bopo_admin/Merchant/add_merchant.html", {"corporates": corporates})







def redirect_with_error(request, message):
    from django.contrib import messages
    messages.error(request, message)
    return redirect("add_merchant")

def redirect_with_success(request, message):
    from django.contrib import messages
    messages.success(request, message)
    return redirect("add_merchant")

# from django.http import JsonResponse
# from accounts.models import Corporate  # or whatever your model is named

# def edit_cop(request, corporate_id):
#     try:
#         merchant = Corporate.objects.get(id=corporate_id)
#         data = {
#             "customer_id": corporate.id,
#             "first_name": corporate.first_name,
#             "last_name": corporate.last_name,
#             "email": corporate.email,
#             "aadhaar": merchant.aadhaar,
#             "shop_name": merchant.shop_name,
#             "address": merchant.address,
#             "state": merchant.state.id if merchant.state else "",
#             "city": merchant.city.id if merchant.city else "",
#             "mobile": merchant.mobile,
#             "gst_number": merchant.gst_number,
#             "pan_number": merchant.pan_number,
#             "legal_name": merchant.legal_name,
#             "pincode": merchant.pincode,
#             "project_name": merchant.project.project_name if merchant.project else "",
#         }
#         return JsonResponse(data)
#     except Corporate.DoesNotExist:
#         return JsonResponse({'error': 'Corporate merchant not found'}, status=404)





# In your Django views.py
# from django.http import JsonResponse
# from accounts.models import Merchant,Corporate


# def edit_copmerchant(request, merchant_id):
#     try:
#         merchant = Merchant.objects.get(id=merchant_id)
#         data = {
#             'id': merchant.merchant_id,
#             'first_name': merchant.first_name,
#             'last_name': merchant.last_name,
#             'email': merchant.email,
#             'aadhaar_number': merchant.aadhaar_number,
#             'shop_name': merchant.shop_name,
#             'address': merchant.address,
#             'state': merchant.state,
#             'city': merchant.city,
#             'pincode': merchant.pincode,
#             'mobile': merchant.mobile,
#             'gst_number': merchant.gst_number,
#             'pan_number': merchant.pan_number,
#             'legal_name': merchant.legal_name,
#             'project_name': merchant.project_name,
#         }
#         return JsonResponse(data)

#     except Merchant.DoesNotExist:
#         return JsonResponse({'error': 'Merchant not found'}, status=404)

def get_corporate(request, corporate_id):
    try:
        corporate = Corporate.objects.get(corporate_id=corporate_id)

        # Assuming state is a string, get the State object by name
        state_obj = State.objects.get(name=corporate.state)

        # Retrieve cities based on selected state
        cities = City.objects.filter(state=state_obj)

        # Convert cities to a dictionary for use in the frontend
        city_data = [{"id": city.id, "name": city.name} for city in cities]


        data = {
            'corporate_id': corporate.corporate_id,
            'project_name': corporate.project_name,
            # 'project_id': corporate.project_id,
            'first_name': corporate.first_name,
            'last_name': corporate.last_name,
            'email': corporate.email,
            'mobile': corporate.mobile,
            'aadhaar_number': corporate.aadhaar_number,
            'pin': corporate.pin,
            'gst_number': corporate.gst_number,
            'pan_number': corporate.pan_number,
            'shop_name': corporate.shop_name,
            'legal_name': corporate.legal_name,
            'address': corporate.address,
            'pincode': corporate.pincode,
            "state": corporate.state,  
            "city": corporate.city,   
            'country': corporate.country,
            "states": [{"id": state.id, "name": state.name} for state in State.objects.all()],  
            "cities": city_data,  
        }

        return JsonResponse(data)

    except Corporate.DoesNotExist:
        return JsonResponse({'error': 'Corporate not found'}, status=404)
    except State.DoesNotExist:
        return JsonResponse({'error': 'State not found'}, status=404)
    except City.DoesNotExist:
        return JsonResponse({'error': 'City not found'}, status=404)


from django.http import JsonResponse
from accounts.models import Corporate


def update_corporate(request):
    if request.method == 'POST':
        corporate_id = request.POST.get('corporate_id')
        try:
            corporate = Corporate.objects.get(corporate_id=corporate_id)

            corporate.first_name = request.POST.get('first_name', '').strip()
            corporate.last_name = request.POST.get('last_name', '').strip()
            corporate.email = request.POST.get('email', '').strip()
            corporate.mobile = request.POST.get('mobile', '').strip()
            corporate.aadhaar_number = request.POST.get('aadhaar_number', '').strip()
            corporate.pin = request.POST.get('pin', '').strip()
            corporate.pan_number = request.POST.get('pan_number', '').strip()
            corporate.gst_number = request.POST.get('gst_number', '').strip()
            corporate.legal_name = request.POST.get('legal_name', '').strip()
            corporate.shop_name = request.POST.get('shop_name', '').strip()
            corporate.address = request.POST.get('address', '').strip()
            corporate.pincode = request.POST.get('pincode', '').strip()
            corporate.project_name = request.POST.get('project_name', '').strip()
            corporate.country = request.POST.get('country', 'India').strip()

            state_id = request.POST.get('state')
            city_id = request.POST.get('city')

            if state_id:
                state = State.objects.get(id=state_id)
                corporate.state = state.name

            if city_id:
                city = City.objects.get(id=city_id)
                corporate.city = city.name

            corporate.save()

            return JsonResponse({
                'success': True,
                'message': 'Corporate updated successfully!',
                'updatedCorporate': {
                    'corporate_id': corporate.corporate_id,
                    'first_name': corporate.first_name,
                    'last_name': corporate.last_name,
                    'email': corporate.email,
                    'mobile': corporate.mobile,
                    'state': corporate.state,
                    'city': corporate.city,
                }
            })
        except Corporate.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Corporate not found'}, status=404)
        except State.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'State not found'}, status=404)
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'City not found'}, status=404)

    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)



from django.http import JsonResponse
from accounts.models import Merchant

def get_copmerchant(request, merchant_id):
    try:
        merchant = Merchant.objects.get(id=merchant_id)

        state_obj = State.objects.get(name=merchant.state)
        city_obj = City.objects.get(name=merchant.city)

        cities = City.objects.filter(state=state_obj)
        city_data = [{"id": city.id, "name": city.name} for city in cities]

        data = {
            'merchant_id': merchant.id,
            'first_name': merchant.first_name,
            'last_name': merchant.last_name,
            'email': merchant.email,
            'mobile': merchant.mobile,
            'aadhaar_number': merchant.aadhaar_number,
            'pan_number': merchant.pan_number,
            'gst_number': merchant.gst_number,
            'legal_name': merchant.legal_name,
            'project_name': merchant.project_name.project_name if merchant.project_name else None,  # ✅ FIXED
            'shop_name': merchant.shop_name,
            'address': merchant.address,
            'pincode': merchant.pincode,
            "state": merchant.state,
            "city": merchant.city,
            'country': merchant.country,
            "states": [{"id": state.id, "name": state.name} for state in State.objects.all()],
            "cities": city_data,
        }

        return JsonResponse(data)

    except Merchant.DoesNotExist:
        return JsonResponse({'error': 'Merchant not found'}, status=404)
    except State.DoesNotExist:
        return JsonResponse({'error': 'State not found'}, status=404)
    except City.DoesNotExist:
        return JsonResponse({'error': 'City not found'}, status=404)



from django.http import JsonResponse
from accounts.models import Merchant, Corporate


def update_copmerchant(request):
    if request.method == "POST":
        merchant_id = request.POST.get('merchant_id')
        if not merchant_id:
            return JsonResponse({'success': False, 'error': 'Missing merchant ID'}, status=400)

        try:
            merchant = Merchant.objects.get(id=merchant_id)

            # Sanitize and update simple fields
            merchant.first_name = request.POST.get('first_name', '').strip()
            merchant.last_name = request.POST.get('last_name', '').strip()
            merchant.email = request.POST.get('email', '').strip()
            merchant.mobile = request.POST.get('mobile', '').strip()
            merchant.aadhaar_number = request.POST.get('aadhaar_number', '').strip()
            merchant.pin = request.POST.get('pin', '').strip()
            merchant.pan_number = request.POST.get('pan_number', '').strip()
            merchant.gst_number = request.POST.get('gst_number', '').strip()
            merchant.legal_name = request.POST.get('legal_name', '').strip()
            merchant.shop_name = request.POST.get('shop_name', '').strip()
            merchant.address = request.POST.get('address', '').strip()
            merchant.pincode = request.POST.get('pincode', '').strip()
            merchant.country = request.POST.get("country", "India").strip()

            # Resolve and set foreign key fields using IDs
            state_id = request.POST.get('state')
            city_id = request.POST.get('city')
            if state_id:
                try:
                    state = State.objects.get(id=state_id)
                    merchant.state = state.name
                except State.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'State not found'}, status=404)
            if city_id:
                try:
                    city = City.objects.get(id=city_id)
                    merchant.city = city.name
                except City.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'City not found'}, status=404)

            # Resolve project_name field (assumed to be a foreign key to Corporate)
            # Here we assume the submitted value is the corporate_id of the project
            project_identifier = request.POST.get('project_name', '').strip()
            if project_identifier:
                try:
                    corporate_obj = Corporate.objects.get(corporate_id=project_identifier)
                    merchant.project_name = corporate_obj
                except Corporate.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Corporate (project) not found'}, status=404)

            merchant.save()

            return JsonResponse({
                "success": True,
                "message": "Merchant updated successfully!",
                "updatedMerchant": {
                    "id": merchant.id,
                    "first_name": merchant.first_name,
                    "last_name": merchant.last_name,
                    "email": merchant.email,
                    "mobile": merchant.mobile,
                    "state": merchant.state,
                    "city": merchant.city
                }
            })

        except Merchant.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Merchant not found'}, status=404)

    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)



from django.http import JsonResponse
from accounts.models import Merchant

def get_copmerchant(request, merchant_id):
    try:
        merchant = Merchant.objects.get(id=merchant_id)

        state_obj = State.objects.get(name=merchant.state)
        city_obj = City.objects.get(name=merchant.city)

        cities = City.objects.filter(state=state_obj)
        city_data = [{"id": city.id, "name": city.name} for city in cities]

        data = {
            'merchant_id': merchant.id,
            'first_name': merchant.first_name,
            'last_name': merchant.last_name,
            'email': merchant.email,
            'mobile': merchant.mobile,
            'aadhaar_number': merchant.aadhaar_number,
            'pin': merchant.pin,
            'pan_number': merchant.pan_number,
            'gst_number': merchant.gst_number,
            'legal_name': merchant.legal_name,
            'project_name': merchant.project_name.project_name if merchant.project_name else None,  # ✅ FIXED
            'shop_name': merchant.shop_name,
            'address': merchant.address,
            'pincode': merchant.pincode,
            "state": merchant.state,
            "city": merchant.city,
            'country': merchant.country,
            "states": [{"id": state.id, "name": state.name} for state in State.objects.all()],
            "cities": city_data,
        }

        return JsonResponse(data)

    except Merchant.DoesNotExist:
        return JsonResponse({'error': 'Merchant not found'}, status=404)
    except State.DoesNotExist:
        return JsonResponse({'error': 'State not found'}, status=404)
    except City.DoesNotExist:
        return JsonResponse({'error': 'City not found'}, status=404)



# from django.http import JsonResponse
# from accounts.models import Merchant, Corporate

# def update_copmerchant(request):
#     if request.method == 'POST':
#         # Extract the data from the request (Assuming data comes as form data)
#         merchant_id = request.POST.get('merchant_id')
#         first_name = request.POST.get('first_name')
#         last_name = request.POST.get('last_name')
#         email = request.POST.get('email')
#         aadhaar = request.POST.get('aadhaar')
#         shop_name = request.POST.get('shop_name')
#         address = request.POST.get('address')
#         state = request.POST.get('state')
#         mobile = request.POST.get('mobile')
#         gst_number = request.POST.get('gst_number')
#         pan_number = request.POST.get('pan')
#         legal_name = request.POST.get('legal_name')
#         city = request.POST.get('city')
#         pincode = request.POST.get('pincode')
#         project_name = request.POST.get('project_name')
#         # select_project = request.POST.get('select_project')

#         try:
#             # Retrieve the merchant to update
#             merchant = Merchant.objects.get(id=merchant_id)

#             # Update the merchant fields
#             merchant.first_name = first_name
#             merchant.last_name = last_name
#             merchant.email = email
#             merchant.aadhaar_number = aadhaar
#             merchant.shop_name = shop_name
#             merchant.address = address
#             merchant.state = state
#             merchant.mobile = mobile
#             merchant.gst_number = gst_number
#             merchant.pan_number = pan_number
#             merchant.legal_name = legal_name
#             merchant.city = city
#             merchant.pincode = pincode
#             merchant.project_name = project_name
#             # merchant.select_project = select_project

#             # Save the updated merchant
#             merchant.save()

#             # Return a success response with the updated merchant data
#             return JsonResponse({
#                 'success': True 
#                 })

#         except Merchant.DoesNotExist:
#             return JsonResponse({'success': False, 'message': 'Merchant not found.'})

#     else:
#         return JsonResponse({'success': False, 'message': 'Invalid request method.'})


from django.http import JsonResponse
from accounts.models import Corporate, Merchant

def delete_corporate(request, id):
    if request.method == 'DELETE':
        try:
            Corporate.objects.get(id=id).delete()
            return JsonResponse({'success': True})
        except Corporate.DoesNotExist:
            return JsonResponse({'error': 'Corporate not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

# def edit_individual(request, id):
#     merchant = get_object_or_404(Merchant, id=id)
#     return render(request, 'bopo_admin/Merchant/edit_individual.html', {'merchant': merchant})



from django.http import JsonResponse
from accounts.models import Merchant

def edit_merchants(request, merchant_id):
    merchant = get_object_or_404(Merchant, id=merchant_id)

    # Retrieve the state object by its name
    state_obj = State.objects.get(name=merchant.state)  # Assuming state is a string, get State object by name

    # Retrieve cities based on selected state
    cities = City.objects.filter(state=state_obj)  # Now we use the State object

    # Convert cities to a dictionary for use in the frontend
    city_data = [{"id": city.id, "name": city.name} for city in cities]
    

    # Data to send to the frontend
    data = {
        "id": merchant.id,
        "first_name": merchant.first_name,
        "last_name": merchant.last_name,
        "email": merchant.email,
        "mobile": merchant.mobile,
        "shop_name": merchant.shop_name,
        "address": merchant.address,
        "aadhaar_number": merchant.aadhaar_number,
        "pin": merchant.pin,
        "gst_number": merchant.gst_number,
        "pan_number": merchant.pan_number,
        "legal_name": merchant.legal_name,
        "state": merchant.state,
        "city": merchant.city,
        "pincode": merchant.pincode,
    }
    return JsonResponse(data)


from django.http import JsonResponse
from accounts.models import Merchant

def update_merchant(request): 
    if request.method == "POST":
        merchant_id = request.POST.get('merchant_id')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        aadhaar = request.POST.get("aadhaar")
        pin = request.POST.get("pin")
        address = request.POST.get("address")
        state_id = request.POST.get("state")
        city_id = request.POST.get("city")
        mobile = request.POST.get("mobile")
        pan = request.POST.get("pan")
        pincode = request.POST.get("pincode")
        shop_name = request.POST.get("shop_name")
        country = request.POST.get("country", "India")
        select_state = request.POST.get("select_state")
        
        try:
            # Get the  merchant object by id
            merchant = Merchant.objects.get(id=merchant_id)
            
            # Update the  merchant object fields
            merchant.first_name= first_name
            merchant.last_name=last_name
            merchant.email = email
            merchant.aadhaar = aadhaar  # Corrected here
            merchant.pin = pin  # Corrected here
            merchant.address = address  # Corrected here
            merchant.state_id = state_id  # Corrected here
            merchant.city_id = city_id  # Corrected here
            merchant.mobile = mobile
            merchant.pan = pan
            merchant.pincode = pincode
            merchant.shop_name = shop_name
            merchant.country = country
            select_state=select_state,
            
            # Save the updated merchant object
            merchant.save()

            return JsonResponse({
            "success": True,
            "message": "Merchant updated successfully!"  # ✅ Important!
        })
        except Merchant.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Merchant not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

from django.http import JsonResponse
from accounts.models import Merchant  # change this based on your model name


def delete_merchant(request, merchant_id):
    if request.method == "DELETE":
        try:
            merchant = Merchant.objects.get(id=merchant_id)
            merchant.delete()
            return JsonResponse({"success": True, "message": "Merchant deleted successfully."})
        except Merchant.DoesNotExist:
            return JsonResponse({"success": False, "message": "Merchant not found."})
    return JsonResponse({"success": False, "message": "Invalid request method."})



from accounts.models import Merchant
from django.shortcuts import render, redirect
from django.http import JsonResponse


def add_individual_merchant(request):
    if request.method == "POST":
        try:
            # Get form data
            first_name = request.POST.get("first_name")
            last_name = request.POST.get("last_name")
            email = request.POST.get("email")
            mobile = request.POST.get("mobile")
            aadhaar_number = request.POST.get("aadhaar_number")
            pin = request.POST.get("pin")
            gst_number = request.POST.get("gst_number")
            pan_number = request.POST.get("pan_number")
            shop_name = request.POST.get("shop_name")
            legal_name = request.POST.get("legal_name")
            address = request.POST.get("address")
            pincode = request.POST.get("pincode")
            state_id = request.POST.get("state")
            city_id = request.POST.get("city")
            country = request.POST.get("country", "India")

            state = State.objects.get(id=state_id)
            city = City.objects.get(id=city_id)

            # Uniqueness checks
            if Merchant.objects.filter(email=email).exists() or Corporate.objects.filter(email=email).exists():
                return JsonResponse({"success": False, "message": "Email ID already exists!"})
            if Merchant.objects.filter(mobile=mobile).exists():
                return JsonResponse({"success": False, "message": "Mobile number already exists!"})
            if Merchant.objects.filter(aadhaar_number=aadhaar_number).exists():
                return JsonResponse({"success": False, "message": "Aadhaar number already exists!"})
            if Merchant.objects.filter(pan_number=pan_number).exists():
                return JsonResponse({"success": False, "message": "PAN number already exists!"})

            # Generate merchant_id
            last_merchant = Merchant.objects.order_by('-id').first()
            next_id = 1 if not last_merchant else last_merchant.id + 1
            merchant_id = f"MID{str(next_id).zfill(8)}"

            # Generate Terminal ID and TID PIN
            def generate_terminal_id():
                return "TID" + ''.join(random.choices(string.digits, k=8))

            terminal_id = generate_terminal_id()
            while Terminal.objects.filter(terminal_id=terminal_id).exists():
                terminal_id = generate_terminal_id()

            tid_pin = random.randint(1000, 9999)

            # ✅ Create the merchant
            merchant = Merchant.objects.create(
                merchant_id=merchant_id,
                first_name=first_name,
                last_name=last_name,
                email=email,
                mobile=mobile,
                gst_number=gst_number,
                aadhaar_number=aadhaar_number,
                pin = pin,
                pan_number=pan_number,
                shop_name=shop_name,
                legal_name=legal_name,
                address=address,
                pincode=pincode,
                state=state,
                city=city,
                country=country,
            
            )

            # ✅ Save terminal info
            Terminal.objects.create(
                terminal_id=terminal_id,
                tid_pin=tid_pin,
                merchant_id=merchant
            )

            return JsonResponse({'success': True, 'message': 'Merchant added successfully!', 'merchant_id': merchant_id})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f"An error occurred: {str(e)}"})

    return render(request, "bopo_admin/Merchant/add_individual_merchant.html")



def project_onboarding(request):
    return render(request, 'bopo_admin/project_onboarding.html')

def project_list(request):
    return render(request, 'bopo_admin/project_list.html')

def merchant_credentials(request):
    merchants = Merchant.objects.all().order_by('merchant_id')
    corporates = Corporate.objects.all().order_by('project_name')

    if request.method == 'POST':
        project_id = request.POST.get('project')
        merchant_id = request.POST.get('merchant_id')
        terminal_id = request.POST.get('terminal_id_dropdown')

        try:
            merchant = Merchant.objects.get(merchant_id=merchant_id)
            phone_number = merchant.mobile
            if not phone_number.startswith('+'):
                phone_number = f'+91{phone_number}'

           # Compose the SMS message
            message_text = (
                f"Dear {merchant.first_name},\n\n"
                f"Your BOPO login credentials are as follows:\n"
                f"Merchant ID : {merchant_id}\n"
                f"Terminal ID : {terminal_id}\n\n"
                f"Please use these credentials to access your BOPO account.\n\n"
                f"Regards,\n"
                f"BOPO Support Team"
            )
           
            # Fetch Twilio credentials from Django settings
            account_sid = settings.TWILIO_ACCOUNT_SID
            auth_token = settings.TWILIO_AUTH_TOKEN
            twilio_phone_number = settings.TWILIO_PHONE_NUMBER

            # Send SMS using Twilio
            client = Client(account_sid, auth_token)
            client.messages.create(
                body=message_text,
                from_=twilio_phone_number,
                to=phone_number
            )
           
            messages.success(request, f"Credentials sent to {merchant.first_name} at {phone_number}")

        except Merchant.DoesNotExist:
            messages.error(request, "Merchant not found.")
        except Exception as e:
            messages.error(request, f"Error sending SMS: {str(e)}")
            print("Sending SMS to:", phone_number)
            print("merchnat id:", merchant_id)

        return redirect('merchant_credentials')

    context = {
        'merchants': merchants,
        'corporates': corporates,
    }
    return render(request, 'bopo_admin/Merchant/merchant_credentials.html', context)

def merchant_topup(request):
    if request.method == "POST":
        merchant_id = request.POST.get("merchant_id")
        topup_points = int(request.POST.get("topup_points") or 0)
        topup_amount = request.POST.get("topup_amount")
        transaction_id = request.POST.get("transaction_id")
        payment_mode = request.POST.get("payment_mode")
        upi_id = request.POST.get("upi_id") or None
        

        try:
            merchant_obj = Merchant.objects.get(merchant_id=merchant_id)

            try:
                # Get latest payment for the merchant
                latest_payment = PaymentDetails.objects.filter(merchant=merchant_obj).latest('id')
                paid_amount = int(latest_payment.paid_amount or 0)

                if int(topup_points) != int(paid_amount):
                    return render(request, 'bopo_admin/Merchant/merchant_topup.html', {
                        "merchants": Merchant.objects.all(),
                        "error": f"Top-Up Points ({topup_points}) must match Paid Amount ({paid_amount}). Transfer blocked."
                    })


            except PaymentDetails.DoesNotExist:
                return render(request, 'bopo_admin/Merchant/merchant_topup.html', {
                    "merchants": Merchant.objects.all(),
                    "error": "No payment details found for the selected merchant."
                })

            # Create Topup entry
            Topup.objects.create(
                merchant=merchant_obj,
                topup_amount=topup_amount,
                transaction_id=transaction_id,
                topup_points=topup_points,
                payment_mode=payment_mode,
                upi_id=upi_id,
                
            )

            # Update or create MerchantPoints
            points_obj, created = MerchantPoints.objects.get_or_create(
                merchant_id=merchant_obj.id,
                defaults={'points': topup_points}
            )
            if not created:
                points_obj.points += topup_points
                points_obj.save()

            return render(request, 'bopo_admin/Merchant/merchant_topup.html', {
                "merchants": Merchant.objects.all(),
                "success": "Top-up successful!"
            })

        except Merchant.DoesNotExist:
            return render(request, 'bopo_admin/Merchant/merchant_topup.html', {
                "merchants": Merchant.objects.all(),
                "error": "Merchant not found"
            })

    return render(request, 'bopo_admin/Merchant/merchant_topup.html', {
        "merchants": Merchant.objects.all()
    })

def get_payment_details(request):
    merchant_code = request.GET.get('merchant_id')

    try:
        merchant_obj = Merchant.objects.get(merchant_id__iexact=merchant_code.strip())
        latest_payment = PaymentDetails.objects.filter(merchant=merchant_obj).latest('id')

        return JsonResponse({
            "paid_amount": latest_payment.paid_amount,
            "transaction_id": latest_payment.transaction_id,
            "payment_mode": latest_payment.payment_mode,
        })

    except Merchant.DoesNotExist:
        return JsonResponse({"error": "Merchant not found"}, status=404)
    except PaymentDetails.DoesNotExist:
        return JsonResponse({"error": "No payment found for this merchant"}, status=404)




def get_merchant_details(request):
    merchant_id = request.GET.get('merchant_id')
    try:
        merchant = Merchant.objects.get(id=merchant_id)
        points_obj = MerchantPoints.objects.get(merchant=merchant)
        return JsonResponse({
            'merchant_code': merchant.merchant_id,
            'points': points_obj.points
        })
    except Merchant.DoesNotExist:
        return JsonResponse({'error': 'Merchant not found'}, status=404)
    except MerchantPoints.DoesNotExist:
        return JsonResponse({
            'merchant_code': merchant.merchant_id,
            'points': 0
        })  


def map_bonus_points(request):
    return render(request, 'bopo_admin/Merchant/map_bonus_points.html')

def merchant_limit_list(request):
    # Fetch all topups along with the merchant related to each topup
    topups = Topup.objects.select_related('merchant').all().order_by('-created_at')

    context = {
        'topups': topups
    }
    return render(request, 'bopo_admin/Merchant/merchant_limit_list.html', context)
    # return render(request, 'bopo_admin/Merchant/merchant_limit_list.html')
    
    
def get_current_limit(request):
    merchant_id = request.GET.get('merchant_id')
    if merchant_id:
        try:
            merchant_points = MerchantPoints.objects.get(merchant_id=merchant_id)
            return JsonResponse({'current_limit': merchant_points.points})
        except MerchantPoints.DoesNotExist:
            return JsonResponse({'current_limit': 0})  # If no record, return 0
    return JsonResponse({'current_limit': 0})
    

def reduce_limit(request):
    corporates = Corporate.objects.all()
    if request.method == "POST":
        project = request.POST.get("project")
        merchant = request.POST.get("merchant")
        current_limit = request.POST.get("current_limit")
        reduce_amount = request.POST.get("reduce_amount")
        transaction_id = request.POST.get("transaction_id")

        print('project:', project)
        print('merchant:', merchant)
        print('current_limit:', current_limit)
        print('reduce limit:', reduce_amount)
        print('transaction id:', transaction_id)

        # Save to database or perform any other action
        Reducelimit.objects.create(
            project=project,
            merchant=merchant,
            current_limit=current_limit,
            reduce_amount=reduce_amount,
            transaction_id=transaction_id
        )
    return render(request, 'bopo_admin/Merchant/reduce_limit.html',  {"corporates": corporates})


def get_merchants_by_project(request):
    project_id = request.GET.get('project_id')
    merchants = Merchant.objects.filter(project_name__project_id=project_id).values('merchant_id', 'first_name', 'last_name')
    return JsonResponse({'merchants': list(merchants)})


def merchant_status(request):
    merchants = Merchant.objects.all().order_by('merchant_id')
    corporates = Corporate.objects.all().order_by('project_name')

    context = {
        'merchants': merchants,
        'corporates': corporates,
        
    }

    return render(request, 'bopo_admin/Merchant/merchant_status.html', context)

def get_terminal_ids(request):
    merchant_code = request.GET.get('merchant_id')  # This is like 'MER000001'
    try:
        merchant = Merchant.objects.get(merchant_id=merchant_code)
        terminal_ids = Terminal.objects.filter(merchant_id=merchant.id).values_list('terminal_id', flat=True)
        return JsonResponse({'terminal_ids': list(terminal_ids)})
    except Merchant.DoesNotExist:
        return JsonResponse({'terminal_ids': []})

def get_merchants(request):
    project_id = request.GET.get('project_id')

    merchants = Merchant.objects.filter(project_id=project_id).values('merchant_id', 'first_name', 'last_name')
    print(list(merchants))
    return JsonResponse({'merchants': list(merchants)})





def login_page_info(request):
    if request.method == 'POST':
        sales_name = request.POST.get('sales_name')
        sales_number = request.POST.get('sales_number')
        sales_email = request.POST.get('sales_email')

        MerchantLogin.objects.create(
            
            sales_name=sales_name,
            sales_number=sales_number,
            sales_email=sales_email
        )
    return render(request, 'bopo_admin/Merchant/login_page_info.html')

def send_sms(to_number, message_body):
    from twilio.rest import Client
    from django.conf import settings

    print("Initializing Twilio Client...")
    print("To:", to_number)
    print("Message:", message_body)
    print("TWILIO_ACCOUNT_SID:", settings.TWILIO_ACCOUNT_SID)
    print("TWILIO_PHONE_NUMBER:", settings.TWILIO_PHONE_NUMBER)

    try:
        # Add +91 if missing and ensure only digits
        if not to_number.startswith("+"):
            if to_number.startswith("0"):
                to_number = to_number[1:]  # remove leading 0
            to_number = "+91" + to_number

        client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
        message = client.messages.create(
            body=message_body,
            from_=settings.TWILIO_PHONE_NUMBER,
            to=to_number
        )
        print("SMS sent. SID:", message.sid)
        return message.sid
    except Exception as e:
        print(f"❌ Twilio SMS error: {str(e)}")
        return None



def send_notifications(request):
    corporates = Corporate.objects.all()

    if request.method == "POST":
        form_type = request.POST.get("form_type")
        project = request.POST.get("project")
        notification_type = request.POST.get("notification_type")
        notification_title = request.POST.get("notification_title")
        description = request.POST.get("description")

        if form_type == "single":
            merchant_id = request.POST.get("merchant")
            merchant = Merchant.objects.get(merchant_id=merchant_id)

            # Save the notification
            Notification.objects.create(
                project_id=project,
                merchant_id=merchant_id,
                notification_type=notification_type,
                title=notification_title,
                description=description
            )

            # Create message and send via SMS
            message = f"{notification_type} - {notification_title}:\n{description}"
            send_sms(merchant.mobile, message)

        elif form_type == "all":
            merchants = Merchant.objects.filter(corporate_id=project)
            for merchant in merchants:
                Notification.objects.create(
                    project_id=project,
                    merchant_id=merchant.merchant_id,
                    notification_type=notification_type,
                    title=notification_title,
                    description=description
                )

                message = f"{notification_type} - {notification_title}:\n{description}"
                send_sms(merchant.mobile, message)

        return render(request, 'bopo_admin/Merchant/send_notifications.html', {
            'corporates': corporates,
            'message': 'Notification(s) sent via SMS!'
        })

    return render(request, 'bopo_admin/Merchant/send_notifications.html', {
        'corporates': corporates
    })



def received_offers(request):
    return render(request, 'bopo_admin/Merchant/received_offers.html')

def uploads(request):
    if request.method == "POST":
        file_type = request.POST.get("file_type")
        uploaded_file = request.FILES.get(file_type)

        if file_type and uploaded_file:
            UploadedFile.objects.create(file_type=file_type, file=uploaded_file)

    # Fetch uploaded files filtered by type
    privacy_policy_file = UploadedFile.objects.filter(file_type="privacy_policy").first()
    terms_conditions_file = UploadedFile.objects.filter(file_type="terms_conditions").first()
    user_guide_file = UploadedFile.objects.filter(file_type="user_guide").first()

    return render(request, 'bopo_admin/Merchant/uploads.html', {
        "privacy_policy_file": privacy_policy_file,
        "terms_conditions_file": terms_conditions_file,
        "user_guide_file": user_guide_file,
    })

def  modify_customer_details(request):
    return render(request, 'bopo_admin/Customer/modify_customer_details.html')

def  send_customer_notifications(request):
    return render(request, 'bopo_admin/Customer/send_customer_notifications.html')

def  customer_uploads(request):
    return render(request, 'bopo_admin/Customer/customer_uploads.html')


def add_customer(request): 
    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        mobile = request.POST.get('mobile')
        age = request.POST.get('age')
        gender = request.POST.get('gender')
        aadhar_number = request.POST.get('aadhaar')
        pin = request.POST.get('pin') 
        pan_number = request.POST.get('pan_number')
        address = request.POST.get('address')
        state_id = request.POST.get('state')
        city_id = request.POST.get('city')
        pincode = request.POST.get('pincode')
        country = request.POST.get("country", "India")

        try:
            state = State.objects.get(id=state_id)
            city = City.objects.get(id=city_id)
        except (State.DoesNotExist, City.DoesNotExist):
            return JsonResponse({"success": False, "message": "Invalid state or city selection."})

        # Validation checks
        if Customer.objects.filter(email=email).exists():
            return JsonResponse({"success": False, "message": "Email ID already exists!"})

        if Customer.objects.filter(mobile=mobile).exists():
            return JsonResponse({"success": False, "message": "Mobile number already exists!"})

        if Customer.objects.filter(aadhar_number=aadhar_number).exists():
            return JsonResponse({"success": False, "message": "Aadhaar number already exists!"})

        if Customer.objects.filter(pan_number=pan_number).exists():
            return JsonResponse({"success": False, "message": "PAN number already exists!"})

        # Save the customer
        Customer.objects.create(
            customer_id=customer_id,
            first_name=first_name,
            last_name=last_name,
            email=email,
            mobile=mobile,
            age=age,
            gender=gender,
            aadhar_number=aadhar_number,
            pin=pin,
            pan_number=pan_number,
            address=address,
            state=state,
            city=city,
            pincode=pincode,
            country=country
        )

        return JsonResponse({"success": True, "message": "Customer added successfully!"})
    
    return render(request, 'bopo_admin/Customer/add_customer.html')


def employee_list(request):
    employees = Employee.objects.all()
    return render(request, 'bopo_admin/Employee/employee_list.html', {'employees': employees})


from django.http import JsonResponse
from .models import Employee 

def get_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    # Retrieve the state object by its name
    state_obj = State.objects.get(name=employee.state)  # Assuming state is a string, get State object by name

    # Retrieve cities based on selected state
    cities = City.objects.filter(state=state_obj)  # Now we use the State object

    # Convert cities to a dictionary for use in the frontend
    city_data = [{"id": city.id, "name": city.name} for city in cities]

    # Data to send to the frontend
    data = {
        "id": employee.id,
        "name": employee.name,
        "email": employee.email,
        "mobile": employee.mobile,
        "address": employee.address,
        "aadhaar": employee.aadhaar,
        "pan": employee.pan,
        "pincode": employee.pincode,
        "state": employee.state,  # Assuming state is a string or related field
        "city": employee.city,    # Assuming city is a string or related field
        "username": employee.username,
        "password": employee.password,
        "states": [{"id": state.id, "name": state.name} for state in State.objects.all()],  # List of all states
        "cities": city_data,  # List of cities filtered by state
    }

    return JsonResponse(data)

from django.http import JsonResponse
from .models import Employee


def update_employee(request): 
    if request.method == "POST":
        employee_id = request.POST.get('employee_id')
        name = request.POST.get('employee_name')
        email = request.POST.get('email')
        aadhaar = request.POST.get("aadhaar")
        address = request.POST.get("address")
        state_id = request.POST.get("state")
        city_id = request.POST.get("city")
        mobile = request.POST.get("mobile")
        pan = request.POST.get("pan")
        pincode = request.POST.get("pincode")
        username = request.POST.get("username")
        password = request.POST.get("password")
        country = request.POST.get("country", "India")

        try:
            employee = Employee.objects.get(id=employee_id)
            state_name = State.objects.get(id=state_id).name
            city_name = City.objects.get(id=city_id).name

            employee.name = name
            employee.email = email
            employee.aadhaar = aadhaar
            employee.address = address
            employee.state = state_name
            employee.city = city_name
            employee.mobile = mobile
            employee.pan = pan
            employee.pincode = pincode
            employee.username = username
            employee.password = password
            employee.country = country

            employee.save()

            return JsonResponse({'status': 'success', 'message': 'Employee updated successfully'})
        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee not found'})
        except (State.DoesNotExist, City.DoesNotExist):
            return JsonResponse({'status': 'error', 'message': 'Invalid state or city'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


from django.http import JsonResponse
from .models import Employee

def delete_employee(request, employee_id):
    try:
        employee = Employee.objects.get(employee_id=employee_id)
        employee.delete()
        return JsonResponse({'success': True})
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)


from django.http import JsonResponse
from django.shortcuts import render
from .models import Employee, State, City

from django.db.models import Max
from django.db import IntegrityError
import re

def add_employee(request):
    if request.method == "POST":
        # Generate employee ID by finding the max employee_id and ensuring uniqueness
        last_employee = Employee.objects.filter(employee_id__startswith="EMP").aggregate(max_id=Max('employee_id'))['max_id']
        
        if last_employee:
            try:
                # Extract the numeric part from the last employee_id (e.g., 'EMP00000001' -> 1)
                last_number = int(re.sub(r'\D', '', last_employee))  # Removing non-numeric characters
            except ValueError:
                last_number = 10000000  # Default to a base number if the format is incorrect
        else:
            last_number = 10000000

        next_number = last_number + 1
        employee_id = f"EMP{next_number:06d}"  # Format employee_id with 8 digits, e.g., 'EMP00000001'

        # Check for uniqueness of the generated employee_id
        while Employee.objects.filter(employee_id=employee_id).exists():
            next_number += 1
            employee_id = f"EMP{next_number:06d}"  # Regenerate employee_id if the current one exists

        # Fetch POST data for employee
        name = request.POST.get("employee_name")
        email = request.POST.get("email")
        aadhaar = request.POST.get("aadhaar")
        address = request.POST.get("address")
        state_id = request.POST.get("state")
        city_id = request.POST.get("city")
        mobile = request.POST.get("mobile")
        pan = request.POST.get("pan")
        pincode = request.POST.get("pincode")
        username = request.POST.get("username")
        password = request.POST.get("password")
        country = request.POST.get("country", "India")

        # Validation checks
        if Employee.objects.filter(email=email).exists():
            return JsonResponse({"success": False, "message": "Email ID already exists!"})
        if Employee.objects.filter(mobile=mobile).exists():
            return JsonResponse({"success": False, "message": "Mobile number already exists!"})
        if Employee.objects.filter(aadhaar=aadhaar).exists():
            return JsonResponse({"success": False, "message": "Aadhaar number already exists!"})
        if Employee.objects.filter(pan=pan).exists():
            return JsonResponse({"success": False, "message": "PAN number already exists!"})

        # Fetch state and city objects
        try:
            state = State.objects.get(id=state_id)
            city = City.objects.get(id=city_id)
        except (State.DoesNotExist, City.DoesNotExist):
            return JsonResponse({"success": False, "message": "Invalid state or city selection."})

        # Create employee record
        try:
            employee = Employee(
                employee_id=employee_id,
                name=name,
                email=email,
                aadhaar=aadhaar,
                address=address,
                state=state.name,
                city=city.name,
                mobile=mobile,
                pan=pan,
                pincode=pincode,
                username=username,
                password=password,
                country=country
            )
            employee.save()
        except IntegrityError as e:
            employee.delete()
            # Handle any integrity errors (shouldn't happen, but just in case)
            return JsonResponse({"success": False, "message": f"Error: {str(e)}"})

        bopo_admin = BopoAdmin(username=username, role="employee", employee=employee)
        bopo_admin.set_password(password)  # Hash the password
        bopo_admin.save()
        return JsonResponse({"success": True, "message": "Employee added successfully!"})

    return render(request, 'bopo_admin/Employee/add_employee.html')



from django.contrib import messages

def assign_employee_role(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        try:
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            messages.error(request, "Employee not found.")
            return redirect('assign_employee_role')  # Redirect back to form

        roles_data = {
            'corporate_merchant': 'Corporate Merchant' in request.POST.getlist('roles'),
            'individual_merchant': 'Individual Merchant' in request.POST.getlist('roles'),
            'merchant_send_credentials': 'Merchant Send Credentials' in request.POST.getlist('roles'),
            'merchant_limit': 'Merchant Limit' in request.POST.getlist('roles'),
            'merchant_login_page_info': 'Merchant Login Page-Info' in request.POST.getlist('roles'),
            'merchant_send_notification': 'Merchant Send Notification' in request.POST.getlist('roles'),
            'merchant_received_offers': 'Merchant Received Offers' in request.POST.getlist('roles'),
            'modify_customer_details': 'Modify Customer Details' in request.POST.getlist('roles'),
            'customer_send_notification': 'Customer Send Notification' in request.POST.getlist('roles'),
            'create_employee': 'Create Employee' in request.POST.getlist('roles'),
            'payment_details': 'Payment Details' in request.POST.getlist('roles'),
            'account_info': 'Account-Info' in request.POST.getlist('roles'),
            'reports': 'Reports' in request.POST.getlist('roles'),
            'deduct_amount': 'Deduct Amount' in request.POST.getlist('roles'),
            'helpdesk_action': 'HelpDesk Action' in request.POST.getlist('roles')
        }

        employee_role, created = EmployeeRole.objects.update_or_create(
            employee=employee,
            defaults=roles_data
        )

        messages.success(request, "Roles successfully assigned.")
        return redirect('employee_list')  # This should be the same view that uses Toastr

    else:
        employees = Employee.objects.all()
        return render(request, 'bopo_admin/Employee/employee_role.html', {'employees': employees})



def payment_details(request):
    topups = PaymentDetails.objects.all().order_by('-created_at')  # or any custom ordering

    if request.method == "POST":
        # Handle any POST data if needed
        pass

    return render(request, 'bopo_admin/Payment/payment_details.html', {
        'topups': topups
    })
    


def account_info(request):
    account = AccountInfo.objects.first()  # Get the first account (modify as per your logic)
    
    if request.method == "POST":
        accountNumber = request.POST.get("accountNumber")
        payableTo = request.POST.get("payableTo")
        bankName = request.POST.get("bankName")
        city = request.POST.get("city")
        accountType = request.POST.get("accountType")
        ifscCode = request.POST.get("ifscCode")
        branchName = request.POST.get("branchName")
        pincode = request.POST.get("pincode")

        if account:
            # Update existing record
            account.accountNumber = accountNumber
            account.payableTo = payableTo
            account.bankName = bankName
            account.city = city
            accountType = accountType
            account.ifscCode = ifscCode
            account.branchName = branchName
            account.pincode = pincode
            account.save()
        else:
            # Create new record if none exists
            AccountInfo.objects.create(
                accountNumber=accountNumber,
                payableTo=payableTo,
                bankName=bankName,
                city=city,
                accountType=accountType,
                ifscCode=ifscCode,
                branchName=branchName,
                pincode=pincode
            )

        return JsonResponse({"message": "Account saved successfully!"})

    return render(request, "bopo_admin/Payment/account_info.html", {"account": account})

def reports(request):
    return render(request, 'bopo_admin/Payment/reports.html')

def login_view(request):
    if request.method == 'POST': 
        username = request.POST.get('username')
        password = request.POST.get('password')
        user_type = request.POST.get('user_type')

        user = authenticate(request, username=username, password=password)
        if user_type == "employee":
            role_permissions = EmployeeRole.objects.filter(employee=user.employee)
            if not role_permissions.exists():
                error_message = "You do not have permission to access this page."
                return render(request, 'bopo_admin/login.html', {'error_message': error_message})
        if user:
            login(request, user)
            return redirect('home')
        else:
            error_message = "Invalid credentials"
            return render(request, 'bopo_admin/login.html', {'error_message': error_message})

        # try:
        #     # Filter user by username and user_type if you are storing type
        #     if user_type == "corporate_admin":
        #         user = Corporate.objects.get(username=username)
        #     else:
        #         user = BopoAdmin.objects.get(username=username)
            
        #     # Optional: filter by role too if role is stored
        #     # user = BopoAdmin.objects.get(username=username, role=user_type)

        #     if check_password(password, user.password):
        #         request.session['admin_id'] = user.id
        #         request.session['user_type'] = user_type  # Store the role in session
        #         return redirect('home')
        #     else:
        #         error_message = "Incorrect password"
        # except BopoAdmin.DoesNotExist:
        #     error_message = "User does not exist"

        # return render(request, 'bopo_admin/login.html', {'error_message': error_message})
    
    # GET request
    return render(request, 'bopo_admin/login.html')



# from datetime import timedelta
# from django.utils import timezone

# def login(request):
#     if request.method == 'POST': 
#         username = request.POST.get('username')
#         password = request.POST.get('password')
#         user_type = request.POST.get('user_type')
#         remember_me = request.POST.get('remember_me')  # Capture remember_me checkbox

#         try:
#             user = BopoAdmin.objects.get(username=username)

#             if check_password(password, user.password):
#                 request.session['admin_id'] = user.id
#                 request.session['user_type'] = user_type

#                 if remember_me:
#                     # Set session to expire in 7 days
#                     request.session.set_expiry(7 * 24 * 60 * 60)  # 7 days in seconds
#                 else:
#                     # Session expires when browser is closed
#                     request.session.set_expiry(0)

#                 return redirect('home')
#             else:
#                 error_message = "Incorrect password"
#         except BopoAdmin.DoesNotExist:
#             error_message = "User does not exist"

#         return render(request, 'bopo_admin/login.html', {'error_message': error_message})

#     return render(request, 'bopo_admin/login.html')


  
def export_projects(request):
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Corporate Projects"

    # Add headers to the sheet
    headers = [
        "Corporate ID", "Project ID", "Project Name", "First Name", "Last Name",
        "Email", "Mobile", "Aadhaar", "GST Number", "PAN", "Shop Name",
        "Address", "City", "State", "Country", "Pincode", "Created At"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font =  Font(bold=True)

    # Fetch data from the Corporate table
    corporates = Corporate.objects.all()
    print("Total corporates:", corporates.count())

    if not corporates.exists():
        print("No data found in the Corporate table.")  # Debugging log

    # Add data to the Excel sheet
    for row_num, corporate in enumerate(corporates, 2):
        sheet.cell(row=row_num, column=1, value=corporate.corporate_id or "")
        sheet.cell(row=row_num, column=2, value=corporate.project_id or "")
        sheet.cell(row=row_num, column=3, value=corporate.project_name or "")
        sheet.cell(row=row_num, column=4, value=corporate.first_name or "")
        sheet.cell(row=row_num, column=5, value=corporate.last_name or "")
        sheet.cell(row=row_num, column=6, value=corporate.email or "")
        sheet.cell(row=row_num, column=7, value=corporate.mobile or "")
        sheet.cell(row=row_num, column=8, value=corporate.aadhaar or "")
        sheet.cell(row=row_num, column=9, value=corporate.gst_number or "")
        sheet.cell(row=row_num, column=10, value=corporate.pan_number or "")
        sheet.cell(row=row_num, column=11, value=corporate.shop_name or "")
        sheet.cell(row=row_num, column=12, value=corporate.address or "")
        sheet.cell(row=row_num, column=13, value=corporate.city or "")
        sheet.cell(row=row_num, column=14, value=corporate.state or "")
        sheet.cell(row=row_num, column=15, value=corporate.country or "")
        sheet.cell(row=row_num, column=16, value=corporate.pincode or "")
        sheet.cell(row=row_num, column=17, value=corporate.created_at.strftime("%Y-%m-%d %H:%M:%S") if corporate.created_at else "")

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to download the file
    response = HttpResponse(
        content=buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Corporate_Projects.xlsx"'

    return response
 

def export_merchants(request):
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Merchant Details"

    # Add headers to the sheet
    headers = [
        "Merchant ID", "User Type", "Project Name", "First Name", "Last Name",
        "Email", "Mobile", "Aadhaar", "GST Number", "PAN", "Shop Name",
        "Address", "City", "State", "Country", "Pincode", "Created At"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Fetch data from the Merchant table
    merchants = Merchant.objects.all()
    if not merchants.exists():
        print("No data found in the Merchant table.")  # Debugging log

    # Add data to the Excel sheet
    for row_num, merchant in enumerate(merchants, 2):
        sheet.cell(row=row_num, column=1, value=merchant.merchant_id or "")
        sheet.cell(row=row_num, column=2, value=merchant.user_type or "")
        sheet.cell(row=row_num, column=3, value=merchant.project_name or "")
        sheet.cell(row=row_num, column=4, value=merchant.first_name or "")
        sheet.cell(row=row_num, column=5, value=merchant.last_name or "")
        sheet.cell(row=row_num, column=6, value=merchant.email or "")
        sheet.cell(row=row_num, column=7, value=merchant.mobile or "")
        sheet.cell(row=row_num, column=8, value=merchant.aadhaar_number or "")
        sheet.cell(row=row_num, column=9, value=merchant.gst_number or "")
        sheet.cell(row=row_num, column=10, value=merchant.pan_number or "")
        sheet.cell(row=row_num, column=11, value=merchant.shop_name or "")
        sheet.cell(row=row_num, column=12, value=merchant.address or "")
        sheet.cell(row=row_num, column=13, value=merchant.city or "")
        sheet.cell(row=row_num, column=14, value=merchant.state or "")
        sheet.cell(row=row_num, column=15, value=merchant.country or "")
        sheet.cell(row=row_num, column=16, value=merchant.pincode or "")
        sheet.cell(row=row_num, column=17, value=merchant.created_at.strftime("%Y-%m-%d %H:%M:%S") if merchant.created_at else "")

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to download the file
    response = HttpResponse(
        content=buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Merchants_Projects.xlsx"'

    return response


def export_disabled_merchants(request):
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Disabled Merchants"

    # Add headers to the sheet
    headers = [
        "Merchant ID", "User Type", "Project Name", "First Name", "Last Name",
        "Email", "Mobile", "Aadhaar", "GST Number", "PAN", "Shop Name",
        "Address", "City", "State", "Country", "Pincode", "Status", "Created At"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Fetch data for merchants with status "Inactive"
    disabled_merchants = Merchant.objects.filter(status="Inactive")
    if not disabled_merchants.exists():
        print("No disabled merchants found.")  # Debugging log

    # Add data to the Excel sheet
    for row_num, merchant in enumerate(disabled_merchants, 2):
        sheet.cell(row=row_num, column=1, value=merchant.merchant_id or "")
        sheet.cell(row=row_num, column=2, value=merchant.user_type or "")
        sheet.cell(row=row_num, column=3, value=merchant.project_name or "")
        sheet.cell(row=row_num, column=4, value=merchant.first_name or "")
        sheet.cell(row=row_num, column=5, value=merchant.last_name or "")
        sheet.cell(row=row_num, column=6, value=merchant.email or "")
        sheet.cell(row=row_num, column=7, value=merchant.mobile or "")
        sheet.cell(row=row_num, column=8, value=merchant.aadhaar_number or "")
        sheet.cell(row=row_num, column=9, value=merchant.gst_number or "")
        sheet.cell(row=row_num, column=10, value=merchant.pan_number or "")
        sheet.cell(row=row_num, column=11, value=merchant.shop_name or "")
        sheet.cell(row=row_num, column=12, value=merchant.address or "")
        sheet.cell(row=row_num, column=13, value=merchant.city or "")
        sheet.cell(row=row_num, column=14, value=merchant.state or "")
        sheet.cell(row=row_num, column=15, value=merchant.country or "")
        sheet.cell(row=row_num, column=16, value=merchant.pincode or "")
        sheet.cell(row=row_num, column=17, value=merchant.status or "")
        sheet.cell(row=row_num, column=18, value=merchant.created_at.strftime("%Y-%m-%d %H:%M:%S") if merchant.created_at else "")

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to download the file
    response = HttpResponse(
        content=buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Disabled_Merchants.xlsx"'

    return response

def export_project_wise_balance(request):
    return render(request, 'bopo_admin/Payment/reports.html') 

def export_merchant_wise_balance(request):
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Merchant-Wise Balance"

    # Add headers to the sheet
    headers = [
        "Merchant ID", "Merchant Name", "Email", "Mobile", 
        "Available Points", "Status", "Created At"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Fetch data from the MerchantPoints table
    merchant_points = MerchantPoints.objects.select_related('merchant').all()
    if not merchant_points.exists():
        print("No merchant points found.")  # Debugging log

    # Add data to the Excel sheet
    for row_num, merchant_point in enumerate(merchant_points, 2):
        # total_points = getattr(merchant_point, "total_points", 0)
        # used_points = getattr(merchant_point, "used_points", 0)
        available_points = merchant_point.points

        merchant = merchant_point.merchant
        sheet.cell(row=row_num, column=1, value=merchant.merchant_id or "")
        sheet.cell(row=row_num, column=2, value=f"{merchant.first_name} {merchant.last_name}" or "")
        sheet.cell(row=row_num, column=3, value=merchant.email or "")
        sheet.cell(row=row_num, column=4, value=merchant.mobile or "")
        # sheet.cell(row=row_num, column=5, value=merchant.project_name or "")
        # sheet.cell(row=row_num, column=6, value=total_points)
        # sheet.cell(row=row_num, column=7, value=used_points)
        sheet.cell(row=row_num, column=5, value=available_points)
        sheet.cell(row=row_num, column=6, value=merchant.status or "")
        sheet.cell(row=row_num, column=7, value=merchant.created_at.strftime("%Y-%m-%d %H:%M:%S") if merchant.created_at else "")

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to download the file
    response = HttpResponse(
        content=buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Merchant_Wise_Balance.xlsx"'

    return response


def export_customer_wise_balance(request):
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Customer-Wise Balance"

    # Add headers to the sheet
    headers = [
        "Customer ID", "Customer Name", "Email", "Mobile", 
        "Available Points", "Created At"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Fetch data from the CustomerPoints table
    customer_points = CustomerPoints.objects.select_related('customer').all()
    if not customer_points.exists():
        print("No customer points found.")  # Debugging log

    # Add data to the Excel sheet
    for row_num, customer_point in enumerate(customer_points, 2):
        available_points = customer_point.points

        customer = customer_point.customer
        sheet.cell(row=row_num, column=1, value=customer.customer_id or "")
        sheet.cell(row=row_num, column=2, value=f"{customer.first_name} {customer.last_name}" or "")
        sheet.cell(row=row_num, column=3, value=customer.email or "")
        sheet.cell(row=row_num, column=4, value=customer.mobile or "")
        sheet.cell(row=row_num, column=5, value=available_points)
        # sheet.cell(row=row_num, column=6, value=customer.status or "")
        sheet.cell(row=row_num, column=7, value=customer.created_at.strftime("%Y-%m-%d %H:%M:%S") if customer.created_at else "")

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to download the file
    response = HttpResponse(
        content=buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Customer_wise_Balance.xlsx"'

    return response


    
def export_customer_transaction(request):
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Customer Transaction History"

    # Add headers to the sheet
    headers = [
        "Customer ID", "Customer Name", "Email", "Mobile", 
        "Transaction Type", "Points"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Fetch data from the History table
    transactions = History.objects.select_related('customer').all()
    if not transactions.exists():
        print("No transaction history found.")  # Debugging log

    # Add data to the Excel sheet
    for row_num, transaction in enumerate(transactions, 2):
        customer = transaction.customer
        # sheet.cell(row=row_num, column=1, value=transaction.id or "")
        sheet.cell(row=row_num, column=1, value=customer.customer_id or "")
        sheet.cell(row=row_num, column=2, value=f"{customer.first_name} {customer.last_name}" or "")
        sheet.cell(row=row_num, column=3, value=customer.email or "")
        sheet.cell(row=row_num, column=4, value=customer.mobile or "")
        sheet.cell(row=row_num, column=5, value=transaction.transaction_type or "")
        sheet.cell(row=row_num, column=6, value=transaction.points or 0)
        # sheet.cell(row=row_num, column=8, value=transaction.transaction_date.strftime("%Y-%m-%d %H:%M:%S") if transaction.transaction_date else "")
        # sheet.cell(row=row_num, column=9, value=transaction.description or "")

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to download the file
    response = HttpResponse(
        content=buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Customer_Transaction_History.xlsx"'

    return response

def export_payment_dues(request):
    return render(request, 'bopo_admin/Payment/reports.html') 

def export_award_transaction(request):
    from openpyxl import Workbook
    from django.http import HttpResponse
    from bopo_award.models import MerchantPoints

    # Create an Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Award Transactions"

    # Add headers to the sheet
    headers = ["Merchant ID", "Merchant Name", "Award Points", "Transaction Date"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Fetch award history from MerchantPoints table
    merchant_points = MerchantPoints.objects.all()

    # Add data to the sheet
    for row_num, merchant_point in enumerate(merchant_points, 2):
        sheet.cell(row=row_num, column=1, value=merchant_point.merchant.merchant_id if merchant_point.merchant else "")
        sheet.cell(row=row_num, column=2, value=f"{merchant_point.merchant.first_name} {merchant_point.merchant.last_name}" if merchant_point.merchant else "")
        sheet.cell(row=row_num, column=3, value=merchant_point.points)
        sheet.cell(row=row_num, column=4, value=merchant_point.updated_at.strftime("%Y-%m-%d %H:%M:%S") if merchant_point.updated_at else "")

    # Save the workbook to a BytesIO buffer
    from io import BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to download the file
    response = HttpResponse(
        content=buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Award_Transactions.xlsx"'

    return response

def export_corporate_merchant(request):
    return render(request, 'bopo_admin/Payment/reports.html')


# Get all states
def get_states(request):
    states = State.objects.all().values('id', 'name')
    return JsonResponse(list(states), safe=False)

# Get cities for a given state
def get_cities(request, state_id):
    cities = City.objects.filter(state_id=state_id).values('id', 'name')
    return JsonResponse(list(cities), safe=False)


def deduct_amount(request):
    if request.method == "POST":
        # Get form data
        user_id = request.POST.get("user_id")
        amount = float(request.POST.get("amount"))

        # Example logic (replace with your model logic)
        # Let's say you have a `UserBalance` model
        from .models import UserBalance
        try:
            user = UserBalance.objects.get(user_id=user_id)
            if user.balance >= amount:
                user.balance -= amount
                user.save()
                messages.success(request, f"₹{amount} deducted successfully!")
            else:
                messages.error(request, "Insufficient balance.")
        except UserBalance.DoesNotExist:
            messages.error(request, "User not found.")

    return render(request, "bopo_admin/Superadmin/deduct_amount.html")

def superadmin_functionality(request):
    return render(request, 'bopo_admin/Superadmin/superadmin_functionality.html')
 

# def security_questions(request):
#     return render(request, 'bopo_admin/Superadmin/security_questions.html')

# def rental_plan(request):
#     return render(request, 'bopo_admin/Superadmin/rental_plan.html')

# def award_points(request):
#     return render(request, 'bopo_admin/Superadmin/award_points.html')


def add_security_question(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        question_text = data.get('question', '').strip()
        if question_text:
            question = SecurityQuestion.objects.create(question=question_text)
            return JsonResponse({'id': question.id, 'question': question.question})
        return JsonResponse({'error': 'Invalid question'}, status=400)
    
def set_deduct_amount(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        deduct_amount = data.get('deduct_amount')

        if deduct_amount is None or deduct_amount < 0:
            return JsonResponse({'error': 'Invalid deduct amount.'}, status=400)

        # Always store in ID=1 (single row)
        setting, created = DeductSetting.objects.get_or_create(id=1)
        setting.deduct_percentage = deduct_amount
        setting.save()

        return JsonResponse({'message': 'Deduct amount updated successfully.', 'deduct_percentage': setting.deduct_percentage})
    
    return JsonResponse({'error': 'Invalid method.'}, status=405)